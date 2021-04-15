import openpyxl
import glob

all_data = {}
for f in glob.glob("*.xlsx"):
    book = openpyxl.open(f, read_only=True)
    sheet = book.active
 #   for row in sheet.iter_rows(max_row=sheet.max_row+1, min_row=1, min_col=5, max_col=sheet.max_column):
 #       for col in row:
 #           print(col.value, end = '/')
 #       print()
  #  print()
  #  print()
    for col in range(4, sheet.max_column, 2):
        if sheet[1][col].value not in all_data:
            all_data[sheet[1][col].value] = [0, 0, 0, 0]

        for row in range(2, sheet.max_row + 1):
            # Увеличиваем количество проектов на 1 для этого сотрудника
            all_data[sheet[1][col].value][2] += 1

            if type(sheet[row][col].value) != int or sheet[row][col].value < 0:
                all_data[sheet[1][col].value][0] += 0
            else:
                all_data[sheet[1][col].value][0] += sheet[row][col].value

            if type(sheet[row][col + 1].value) != int or sheet[row][col + 1].value < 0:
                all_data[sheet[1][col].value][1] += 0
            else:
                all_data[sheet[1][col].value][1] += sheet[row][col + 1].value

            # Вычисляем КПД сотрудника по формуле: КПД =
            if (type(sheet[row][col].value) != int or sheet[row][col].value == 0) and (type(sheet[row][col + 1].value) != int or sheet[row][col + 1].value == 0) and all_data[sheet[1][col].value][3] == 0:
                all_data[sheet[1][col].value][3] = 0
            elif (type(sheet[row][col].value) != int or sheet[row][col].value == 0) and (type(sheet[row][col + 1].value) != int or sheet[row][col + 1].value == 0):
                all_data[sheet[1][col].value][3] += 0
            elif sheet[row][col].value == sheet[row][col + 1].value and all_data[sheet[1][col].value][3] == 0:
                all_data[sheet[1][col].value][3] = 1
            elif (type(sheet[row][col].value) != int or sheet[row][col + 1].value == 0) and all_data[sheet[1][col].value][3] > 0:
                all_data[sheet[1][col].value][3] = (all_data[sheet[1][col].value][3] + 1) / all_data[sheet[1][col].value][2]
            else:
                all_data[sheet[1][col].value][3] = (all_data[sheet[1][col].value][3] + sheet[row][col + 1].value / sheet[row][col].value) / all_data[sheet[1][col].value][2]

            print(all_data)
    print()


print(all_data)

# Избавиться от магических чисел в диапазонах !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!